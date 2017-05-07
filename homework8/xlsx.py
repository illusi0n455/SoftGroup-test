from openpyxl import load_workbook

wb = load_workbook('currencies.xlsx')
ws = wb.active
search = "".lower()
for i in range(1, ws.max_row+1):
    if search in ws["B" + str(i)].value.lower() or search in ws["C" + str(i)].value.lower():
        for cell in ws[i]:
            print(cell.value, end=' ')
        print()
