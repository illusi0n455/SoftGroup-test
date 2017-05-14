from functools import wraps
from flask import Flask, request, session, url_for, redirect, render_template, g, flash
from werkzeug import check_password_hash, generate_password_hash
from werkzeug.routing import BaseConverter
from app.models import db, User, lower
from openpyxl import load_workbook
# from flask_wtf.csrf import CSRFProtect
import threading
import scrapper


# app settings
DEBUG = True
SECRET_KEY = 'you-will-never-guess'

# app initialisation
app = Flask(__name__)
# csrf = CSRFProtect(app)
app.config.from_object(__name__)
app.config.from_envvar('EXAPP_SETTINGS', silent=True)
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:1@localhost:5432/postgres'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# initialize sqlalchemy with context of current app
db.init_app(app)


# regex custom routing converter
class RegexConverter(BaseConverter):
    def __init__(self, url_map, *items):
        super(RegexConverter, self).__init__(url_map)
        self.regex = items[0]


app.url_map.converters['regex'] = RegexConverter


# executed each time before request passes to the view
@app.before_request
def before_request():
    g.user = None
    if 'user_id' in session:
        g.user = User.query.filter_by(user_id=session['user_id']).first()


@app.before_first_request
def before_first_request():
    th = threading.Thread(target=scrapper.run)
    th.daemon = True
    th.start()


# prevent unauthorized access
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if g.user is None:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)

    return decorated_function


def get_user_id(username):
    """Convenience method to look up the id for a username."""
    rv = User.query.filter_by(username=username).first()
    return rv[0] if rv else None


@app.route('/id_<regex("\d+"):id>/')
@login_required
def detail(id):
    """
    Example of regex use in routes.
    Gets first post of author with chosen id.
    """
    wb = load_workbook('currencies.xlsx')
    ws = wb.active
    row = str(int(id) + 1)
    keys = ws['A1':'J1'][0]
    values = ws['A' + row:'J' + row][0]
    currencies = [keys, values]
    return render_template('detail.html', currencies=currencies)


@app.route('/', methods=['GET'])
@login_required
def home_page():
    """Displays topics"""
    if request.method == 'GET':
        wb = load_workbook('currencies.xlsx')
        ws = wb.active
        return render_template('index.html', currencies=ws.rows)


@app.route('/', methods=['POST'])
@login_required
def home_page_post():
    """Render results"""
    if request.method == 'POST':
        currencies = []
        search_query = request.form['text'].lower()
        wb = load_workbook('currencies.xlsx')
        ws = wb.active
        for i in range(1, ws.max_row+1):
            if search_query in ws["B" + str(i)].value.lower() or search_query in ws["C" + str(i)].value.lower():
                currencies.append(ws[i])
    return render_template('results.html', currencies=currencies)


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Logs the user in."""
    if g.user:
        return redirect(url_for('home_page'))
    error = None
    if request.method == 'POST':
        user = User.query.filter(lower(User.username) == lower(request.form['username'])).first()
        if user is None:
            error = 'Invalid username'
        elif not check_password_hash(user.pw_hash,
                                     request.form['password']):
            error = 'Invalid password'
        else:
            flash('You were logged in')
            session['user_id'] = user.user_id
            return redirect(url_for('home_page'))
    return render_template('login.html', error=error)


@app.route('/logout', methods=['GET'])
def logout():
    """Logout user."""
    if session.get('user_id', None):
        # clear user session to prevent recognition of current user
        session.clear()
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Registers the user."""
    if g.user:
        return redirect(url_for('home_page'))
    error = None
    if request.method == 'POST':
        if not request.form['username']:
            error = 'You have to enter a username'
        elif not request.form['email'] or '@' not in request.form['email']:
            error = 'You have to enter a valid email address'
        elif not request.form['password']:
            error = 'You have to enter a password'
        elif request.form['password'] != request.form['password2']:
            error = 'The two passwords do not match'
        elif get_user_id(request.form['username']) is not None:
            error = 'The username is already taken'
        else:
            # create new user
            user_data = (request.form['username'], request.form['email'],
                         generate_password_hash(request.form['password']))
            new_user = User(*user_data)

            db.session.add(new_user)
            db.session.commit()

            flash('You were successfully registered and can login now')

            return redirect(url_for('login'))
    return render_template('register.html', error=error)


if __name__ == '__main__':
    app.debug = DEBUG
    app.run()
